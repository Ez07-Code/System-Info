# report_generator.py

import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_html(data, filename):
    """Genera el reporte en formato HTML a partir de los datos recolectados."""
    style = """... (El mismo <style> que tenías en tu script original) ..."""
   
    style = """
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f4f9; color: #333; margin: 0; padding: 20px; }
        .container { max-width: 900px; margin: auto; background: #fff; padding: 20px; border-radius: 8px; box-shadow: 0 0 15px rgba(0,0,0,0.1); }
        h1 { color: #4a69bd; text-align: center; border-bottom: 2px solid #4a69bd; padding-bottom: 10px; }
        h2 { color: #1e3799; border-bottom: 1px solid #ddd; padding-bottom: 5px; margin-top: 30px; }
        table { width: 100%; border-collapse: collapse; margin-top: 15px; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }
        th { background-color: #f2f2f2; font-weight: 600; }
        tr:nth-child(even) { background-color: #f9f9f9; }
        tr:hover { background-color: #e2e8f0; }
        .info-table td:first-child { font-weight: bold; width: 30%; }
        .footer { text-align: center; margin-top: 30px; font-size: 0.9em; color: #777; }
    </style>
    """
    html = f"<!DOCTYPE html><html lang='es'><head><meta charset='UTF-8'><title>Reporte del Sistema</title>{style}</head><body>"
    html += "<div class='container'><h1>Reporte de Información del Sistema</h1>"
    
    if 'system' in data and 'network' in data:
        html += "<h2>Información General y de Usuario</h2><table class='info-table'>"
        for key, value in {**data['system'], **data['network']}.items(): html += f"<tr><td>{key}</td><td>{value}</td></tr>"
        html += "</table>"
    if 'bios' in data and 'ram' in data:
        html += "<h2>Hardware</h2><table class='info-table'>"
        for key, value in {**data['bios'], **data['ram']}.items(): html += f"<tr><td>{key}</td><td>{value}</td></tr>"
        html += "</table>"
    if 'cpu' in data:
        html += "<h2>Procesador (CPU)</h2><table class='info-table'>"
        for key, value in data['cpu'].items(): html += f"<tr><td>{key}</td><td>{value}</td></tr>"
        html += "</table>"
    if 'disks' in data:
        html += "<h2>Discos de Almacenamiento</h2><table><tr><th>Fabricante</th><th>Capacidad Total</th><th>Número de Serie</th></tr>"
        for disk in data['disks']:
            if "Error" in disk: html += f"<tr><td colspan='3'>{disk['Error']}</td></tr>"
            else: html += f"<tr><td>{disk.get('Fabricante', 'N/A')}</td><td>{disk.get('Capacidad Total', 'N/A')}</td><td>{disk.get('Número de Serie', 'N/A')}</td></tr>"
        html += "</table>"
    if 'os' in data:
        html += "<h2>Sistema Operativo</h2><table class='info-table'>"
        for key, value in data['os'].items(): html += f"<tr><td>{key}</td><td>{value}</td></tr>"
        html += "</table>"
    if 'printers' in data:
        html += "<h2>Impresoras Instaladas</h2><table><tr><th>Nombre</th><th>Controlador</th><th>Puerto</th><th>Predeterminada</th></tr>"
        if not data['printers']: html += "<tr><td colspan='4'>No se encontraron impresoras.</td></tr>"
        else:
            for printer in data['printers']:
                if "Error" in printer: html += f"<tr><td colspan='4'>{printer['Error']}</td></tr>"
                else: html += f"<tr><td>{printer['Nombre']}</td><td>{printer['Controlador']}</td><td>{printer['Puerto']}</td><td>{printer['Default']}</td></tr>"
        html += "</table>"
    if 'software' in data:
        html += "<h2>Programas Instalados</h2><table><tr><th>Nombre</th><th>Versión</th><th>Vendedor</th></tr>"
        if not data['software']: html += "<tr><td colspan='3'>No se encontraron programas.</td></tr>"
        else:
            for sw in data['software']:
                if "Error" in sw: html += f"<tr><td colspan='3'>{sw['Error']}</td></tr>"
                else: html += f"<tr><td>{sw['Nombre']}</td><td>{sw.get('Versión','N/A')}</td><td>{sw.get('Vendedor','N/A')}</td></tr>"
        html += "</table>"

    report_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    html += f"<div class='footer'>Reporte generado el {report_time}</div></div></body></html>"

    with open(filename, "w", encoding="utf-8") as f:
        f.write(html)

def generate_pdf(data, filename):
    doc = SimpleDocTemplate(filename, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('Title', parent=styles['h1'], alignment=1, spaceAfter=20, textColor=colors.darkblue)
    heading_style = ParagraphStyle('Heading2', parent=styles['h2'], spaceBefore=10, spaceAfter=10, textColor=colors.darkslateblue)
    
    story.append(Paragraph("Reporte de Información del Sistema", title_style))

    # Estilo de tabla genérico
    table_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.black),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ])
    
    def create_table(title, section_data):
        if not section_data or 'Error' in section_data: return
        story.append(Paragraph(title, heading_style))
        table_data = list(section_data.items())
        t = Table(table_data, colWidths=[150, 300])
        t.setStyle(table_style)
        story.append(t)
        story.append(Spacer(1, 12))

    def create_list_table(title, headers, section_data):
        if not section_data or "Error" in section_data[0]: return
        story.append(Paragraph(title, heading_style))
        table_data = [headers]
        keys = [h.lower().replace(" ", "_") for h in headers] # Simplificado, asume consistencia
        
        # Mapeo de claves para flexibilidad
        key_map = {
            "nombre": "Nombre", "fabricante": "Fabricante", "capacidad_total": "Capacidad Total", 
            "número_de_serie": "Número de Serie", "controlador": "Controlador", "puerto": "Puerto",
            "predeterminada": "Default", "versión": "Versión", "vendedor": "Vendedor"
        }

        for item in section_data:
            row = [item.get(key_map.get(k, k.capitalize()), "N/A") for k in keys]
            table_data.append(row)
        
        t = Table(table_data)
        t.setStyle(table_style)
        story.append(t)
        story.append(Spacer(1, 12))

    if 'system' in data: create_table("Información del Sistema", data['system'])
    if 'network' in data: create_table("Red y Usuario", data['network'])
    if 'bios' in data: create_table("BIOS", data['bios'])
    if 'cpu' in data: create_table("Procesador (CPU)", data['cpu'])
    if 'ram' in data: create_table("Memoria RAM", data['ram'])
    if 'os' in data: create_table("Sistema Operativo", data['os'])

    if 'disks' in data: create_list_table("Discos", ["Fabricante", "Capacidad Total", "Número de Serie"], data['disks'])
    if 'printers' in data: create_list_table("Impresoras", ["Nombre", "Controlador", "Puerto", "Predeterminada"], data['printers'])
    if 'software' in data: create_list_table("Software Instalado", ["Nombre", "Versión", "Vendedor"], data['software'])

    doc.build(story)

def generate_excel(data, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14, color="004A69BD")

    def write_section(ws, title, section_data, row_cursor):
        if not section_data or "Error" in section_data: return row_cursor
        ws.cell(row=row_cursor, column=1, value=title).font = title_font
        row_cursor += 1
        for key, value in section_data.items():
            ws.cell(row=row_cursor, column=1, value=key).font = header_font
            ws.cell(row=row_cursor, column=2, value=str(value)) # Asegurar que todo sea string
            row_cursor += 1
        return row_cursor + 1 # Deja un espacio

    row = 1
    if 'system' in data: row = write_section(ws, "Información del Sistema", data['system'], row)
    if 'network' in data: row = write_section(ws, "Red y Usuario", data['network'], row)
    if 'bios' in data: row = write_section(ws, "BIOS", data['bios'], row)
    if 'cpu' in data: row = write_section(ws, "Procesador (CPU)", data['cpu'], row)
    if 'ram' in data: row = write_section(ws, "Memoria RAM", data['ram'], row)
    if 'os' in data: row = write_section(ws, "Sistema Operativo", data['os'], row)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50

    def write_list_sheet(wb, title, headers, section_data):
        if not section_data or 'Error' in section_data[0]: return
        ws_list = wb.create_sheet(title)
        for i, header in enumerate(headers, 1):
            ws_list.cell(row=1, column=i, value=header).font = header_font
            ws_list.column_dimensions[chr(64+i)].width = 30 if i > 1 else 50
        
        keys = list(section_data[0].keys()) # Asume que todas las entradas tienen las mismas claves
        for r_idx, item in enumerate(section_data, 2):
            for c_idx, key in enumerate(keys, 1):
                ws_list.cell(row=r_idx, column=c_idx, value=str(item.get(key, 'N/A')))
    
    if 'disks' in data: write_list_sheet(wb, "Discos", list(data['disks'][0].keys()), data['disks'])
    if 'printers' in data: write_list_sheet(wb, "Impresoras", list(data['printers'][0].keys()), data['printers'])
    if 'software' in data: write_list_sheet(wb, "Software", list(data['software'][0].keys()), data['software'])

    wb.save(filename)